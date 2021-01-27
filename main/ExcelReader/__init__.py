from openpyxl import load_workbook
from main.URLRequests.getFileFromURL import get_file_from_url
from io import BytesIO

def getExcelFile(url):
    """Get Openpyxl Workbook Type from URL

    This function takes a URL and returns an Openpyxl Workbook type

    ## Parameters
    ### url : `str`
    The direct download link for the excel file.
    """
    # Loads an excel file and initialises it as an Openpyxl Workbook type
    excelFile = load_workbook(
        BytesIO(get_file_from_url(url)),
        read_only=True,
        data_only=True
    )
    # print("All sheet names {} " .format(excelFile.sheetnames))
    return excelFile
